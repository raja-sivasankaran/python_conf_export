#!/usr/bin/python3

import os
import sys
import openpyxl
import re

print('This is the name of the script:', sys.argv[0])
print('Number of arguments:', len(sys.argv))
print('The arguments are:', str(sys.argv))

if len(sys.argv) != 3:
    print('Please input excel sheet and outfile filename')
    sys.exit()

excel_sheet_filename = str(sys.argv[1])
#excel_sheet_filename = str('Apalis_TK1_1.xlsx')
#excel_sheet_filename = str('Colibri_iMX7_512MB.xlsx')

wb = openpyxl.load_workbook(excel_sheet_filename)
ws1 = wb.active
    
print(excel_sheet_filename)
tmp_str = re.split('[_.]', excel_sheet_filename)
result_file_name = str(sys.argv[2]) #tmp_str[0] + '_' + tmp_str[1] + '_' + tmp_str[2] + str('.conf')

# Open a new text file for writing results.
print('Writing results...' + result_file_name)
resultFile = open(result_file_name, 'w')

resultFile.write('# Toradex ' + tmp_str[0] + ' ' + tmp_str[1] + ' Computer On Module.' + '\n')
resultFile.write('# http://developer.toradex.com/products/' + tmp_str[0].lower() + '-' + tmp_str[1].lower() + '\n\n')

resultFile.write('[board]' + '\n')
resultFile.write('dtfile = /proc/device-tree/model' + '\n')
resultFile.write('model = Toradex ' + tmp_str[0] + ' ' + tmp_str[1] + ' on ' + tmp_str[0] + ' Evaluation Board' + '\n\n')

resultFile.write('[GPIO]' + '\n')
resultFile.write('###' + tmp_str[0] + ' ' + tmp_str[1].upper() + ' SODIMM number to GPIO number mapping' + '\n\n')

#column_port_heading = str()

#if tmp_str[1].lower() == 'imx7':
#    column_port_heading = (tmp_str[1] + '_' + tmp_str[2] + str(' Function')).lower()
#elif tmp_str[1].lower() == 'vf50':
#    column_port_heading = str('VF50 Note2')

column_pin_heading = input('Enter sodimm pin column heading name where only the pin number is filled:')
column_port_heading = input('Enter gpio port information column heading name:')


processor_type_str = str(tmp_str[1])

#processor_type = processor_type_str.find('T')

processor_family= str('NXP')

if processor_type_str.find('TK1') >= 0:
    processor_family = str('TK1')
elif processor_type_str.find('T') >= 0:
    processor_family = str('NVIDIA')
elif processor_type_str.find('iMX7') >= 0:
    processor_family = str('iMX7')
elif processor_type_str.find('iMX6') >= 0:
    processor_family = str('iMX6')

column_port = int(1)
column_pin = int(1)
sodimm_number = int(0)

for tmp in range(1, ws1.max_column):
    column_str = ws1.cell(1, tmp).value
    if column_str.lower() == column_pin_heading.lower():
        column_pin = tmp
    if column_str.lower() == column_port_heading.lower():
        column_port = tmp

pin_name = str()

if tmp_str[0].lower() == str('colibri'):
    pin_name = str('SODIMM_')
elif tmp_str[0].lower() == str('apalis'):
    pin_name = str('MXM3_')

print('Reading rows...')
for row in range(2, ws1.max_row + 1):
        # Each row in the spreadsheet has data for one census tract.

    bank_number = int(1)

    if sodimm_number == ws1.cell(row, column_pin).value:
        continue;

    sodimm_number = ws1.cell(row, column_pin).value
    gpio_str = ws1.cell(row, column_port).value

    #This is less time consuming option, but don't know how to use it
    #s = [int(s) for s in gpioStr.split() if s.isdigit()]

#   print(gpio_str)

#   print(gpio_str_split)

    if processor_family == 'NXP':
        gpio_str_split = re.findall('\d+', gpio_str)
        bank_number = int(gpio_str_split[0])
        offset_number = int(gpio_str_split[1])
        gpio_number = bank_number * 32 + offset_number
    elif processor_family == 'iMX7' or processor_family == 'iMX6':
        gpio_str_split = re.findall('\d+', gpio_str)
        bank_number = int(gpio_str_split[0])
        offset_number = int(gpio_str_split[1])
        gpio_number = (bank_number -1) * 32 + offset_number
#        m = re.match("(?:(?:\w{3})|(?:\-{3}))\d\d\d$", v)

    if processor_family == "NVIDIA":
        gpio_str_split = re.findall(r'-[\w]+', gpio_str)
        gpio_str_split = re.findall(r'[A-Za-z]+', gpio_str_split[0])
        #gpio_str_split = re.findall(r'[.]+', gpio_str_split[0])
        gpio_str_lst = list(gpio_str_split[0])

        gpio_number_str = re.findall('\d+', gpio_str)
        #if gpio_str_lst.__len__() == int(1):
        bank_number = ord(gpio_str_lst[0].upper()) - ord('A')

        if gpio_str_lst.__len__() == int(2):
            bank_number = bank_number + 26

        offset_number = int(gpio_number_str[0])
        gpio_number = bank_number * 8 + offset_number

    elif processor_family == "TK1":
        gpio_str_split = list()
        gpio_str_split = re.findall(r'_[\w]+', gpio_str)

        if not gpio_str_split:
            continue

        gpio_str_split = re.findall(r'[A-Za-z]+', gpio_str_split[0])
        # gpio_str_split = re.findall(r'[.]+', gpio_str_split[0])
        gpio_str_lst = list(gpio_str_split[0])

        gpio_number_str = re.findall('\d+', gpio_str)
        # if gpio_str_lst.__len__() == int(1):

        if gpio_str_lst.__len__() == int(1) or gpio_str_lst.__len__() == int(0):
            continue;

        bank_number = ord(gpio_str_lst[1].upper()) - ord('A')

        if gpio_str_lst.__len__() == int(3):
            bank_number = bank_number + 26

        offset_number = int(gpio_number_str[1])
        gpio_number = bank_number * 8 + offset_number

#   print(bank_number, offset_number, gpio_number)

    gpio_number_str = str(gpio_number)
    sodimm_number_check = str(sodimm_number)
    sodimm_number_final = re.findall(r'\d+', sodimm_number_check)[0]
    sodimm_number_str = pin_name + sodimm_number_final

    resultFile.write(sodimm_number_str + ' = ' + gpio_number_str + '\n')
#   print(sodimm_number_str + '=' + gpio_number_str)

    if(gpio_str_split.__len__() == 4):
        bank_number = int(gpio_str_split[2])
        offset_number = int(gpio_str_split[3])
        gpio_number = (bank_number - 1) * 32 + offset_number
        gpio_number_str = str(gpio_number)
        sodimm_number_str += "#"
        resultFile.write(sodimm_number_str + ' = ' + gpio_number_str + '\n')
#       print(sodimm_number_str + '=' + gpio_number_str)

resultFile.close()
print('Done.')
